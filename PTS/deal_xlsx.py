import copy

import openpyxl
import os
class DEAL_XLSX:
    def __init__(self,path):
        self.xlsx_path = path
        self.rows_data = []
        self.pst_table = []
        self.dvfs = []
        self.dvfs_dict = {}
        self.same_net = []
        self.same_net_dict = {}
        self.all_net = []
        self.all_net_dict = {}
        self.ban_pst_table = []
        self.final_pst_table = []
        self.original_index_dict = {}
        self.final_index_dict = {}

    def read_excel_xlsx(self,sheet_name):
        workbook = openpyxl.load_workbook(self.xlsx_path)
        sheet = workbook[sheet_name]
        # way 1
        for row in sheet.rows:
            # collect current column data
            row_text_list = []
            for cell in row:
                row_text_list.append(cell.value)
                #print(cell.value, "\t", end="")
            #print(row_text_list)
            #print()
            self.rows_data.append(row_text_list)

        # way 2
        # for row in sheet.iter_rows(min_row=2):
        #     print(row[1].value)
        #
        # # way 3
        # for row in sheet.iter_rows(min_row=2, max_row=5):
        #     if row[1].value is None:
        #         continue
        #     print(row[1].value)
    def show_xlxs_content(self):
        for i in range(0, len(self.rows_data)):
            for j in range(0, len(self.rows_data[i])):
                if self.rows_data[i][j] != None:
                    print('{0:<10}'.format(self.rows_data[i][j]), end="")
                else:
                    print('{0:<10}'.format(' '), end="")
            print()

    def get_supply_net_info(self):
        pst_flag = 0
        dfvs_flag = 0
        same_net_flag = 0
        all_net_flag = 0
        ban_pst_flag = 0
        for i in range(0, len(self.rows_data)):
            temp_row = []
            for j in range(0, len(self.rows_data[i])):
                if self.rows_data[i][j] != None:
                    if self.rows_data[i][j] == 'PST_TABLE':
                        pst_flag = 1
                    elif self.rows_data[i][j] == 'DVFS':
                        dfvs_flag = 1
                    elif self.rows_data[i][j] == 'SAME_NET':
                        same_net_flag = 1
                    elif self.rows_data[i][j] == 'ALL_NET':
                        all_net_flag = 1
                    elif self.rows_data[i][j] == 'BAN_PST':
                        ban_pst_flag = 1
                    elif ban_pst_flag == 1:
                        temp_row.append(self.rows_data[i][j])
                    elif all_net_flag == 1:
                        temp_row.append(self.rows_data[i][j])
                    elif same_net_flag == 1:
                        temp_row.append(self.rows_data[i][j])
                    elif dfvs_flag == 1:
                        temp_row.append(self.rows_data[i][j])
                    elif pst_flag == 1:
                        temp_row.append(self.rows_data[i][j])
            if ban_pst_flag == 1:
                if len(temp_row) > 0:
                    self.ban_pst_table.append(temp_row)
            elif all_net_flag == 1:
                if len(temp_row) > 0:
                    self.all_net.append(temp_row)
            elif same_net_flag == 1:
                if len(temp_row) > 0:
                    self.same_net.append(temp_row)
            elif dfvs_flag == 1:
                if len(temp_row) > 0:
                    self.dvfs.append(temp_row)
            elif pst_flag == 1:
                if len(temp_row) > 0:
                    self.pst_table.append(temp_row)

            for i in range(0,len(self.dvfs)):
                key = None
                temp_list = []
                for j in range(0, len(self.dvfs[i])):
                    if j == 0:
                        key = self.dvfs[i][j]
                    else:
                        temp_list.append(self.dvfs[i][j])
                self.dvfs_dict[key] = temp_list

            for i in range(0, len(self.same_net)):
                key = None
                reulst = None
                for j in range(0, len(self.same_net[i])):
                    if j == 0:
                        result = self.same_net[i][j]
                    else:
                        self.same_net_dict[self.same_net[i][j]] = result

            for i in range(0, len(self.all_net)):
                for j in range(0, len(self.all_net[i])):
                    if self.all_net[i][j] not in self.same_net_dict.keys():
                        self.all_net_dict[self.all_net[i][j]] = 'SELF'
                    else:
                        self.all_net_dict[self.all_net[i][j]] = self.same_net_dict[self.all_net[i][j]]

    def gen_final_pst_table(self):
        final_pst_head = []
        last_pst = self.pst_table[1::]
        original_head_list = self.pst_table[0]
        temp_pst = []
        temp_final_pst_table = []
        for i in range(0,len(original_head_list)):
            self.original_index_dict[original_head_list[i]] = i
        #print(last_pst)
        #print(self.original_index_dict)
        key_index =0
        for key in self.all_net_dict.keys():
            final_pst_head.append(key)
            self.final_index_dict[key] = key_index
            key_index = key_index +1

        #print("ss1")
        temp_final_pst_table.append(final_pst_head)

        for key in self.dvfs_dict.keys():
            #print("ss2")
            if len(self.dvfs_dict[key]) == 1:
                continue
            #print("ss3")
            inital_pst = copy.deepcopy(last_pst)
            key_result = self.dvfs_dict[key][1::]
            for i in range(0,len(inital_pst)-1):
                temp_pst.append(inital_pst[i])
            #print("haha {}".format(temp_pst))
            for i in range(0,len(inital_pst)):
                index = self.original_index_dict[key]
                temp_pst_one_line = copy.deepcopy(inital_pst[i])
                for j in range(0,len(key_result)):
                    if temp_pst_one_line[index] !="V_off":
                        temp_pst_one_line[index] = key_result[j]
                        temp_pst.append(temp_pst_one_line)
            #print("hehe {}".format(temp_pst))
            temp_pst.append(inital_pst[len(inital_pst)-1])
            sort_temp_pst = self.remove_duplicate_pst(temp_pst)
            last_pst = sort_temp_pst

        print("dfvs_result: {}".format(len(last_pst)))
        print(last_pst)


        for i in range(0,len(last_pst)):
            temp_final_pst_one_line = []
            temp_pst_one_line_reulst = last_pst[i]
            for j in range(0,len(final_pst_head)):
                if final_pst_head[j] not in self.original_index_dict.keys():
                    index = self.original_index_dict[self.same_net_dict[final_pst_head[j]]]
                else:
                    index = self.original_index_dict[final_pst_head[j]]
                temp_final_pst_one_line.append(temp_pst_one_line_reulst[index])
            if self.is_valid_pst(temp_final_pst_one_line):
                temp_final_pst_table.append(temp_final_pst_one_line)
        print("final pst result: {}".format(len(temp_final_pst_table)-1))
        self.final_pst_table = temp_final_pst_table

    def is_valid_pst(self,one_pst_line):
        find_it = True
        for i in range(0,len(self.ban_pst_table)):
            no_valid_state_dict = {}
            one_ban_pst_len = len(self.ban_pst_table[i])
            middle_index = int(one_ban_pst_len / 2)
            for j in range(0,middle_index):
                no_valid_state_dict[self.ban_pst_table[i][j]] = self.ban_pst_table[i][j+middle_index]
            #print(no_valid_state_dict)
            for key in no_valid_state_dict.keys():
                index = self.final_index_dict[key]
                if no_valid_state_dict[key] == one_pst_line[index]:
                    find_it = True
                else:
                    find_it = False
                    break
            if find_it == True:
                break
        if find_it == True:
            return False
        else:
            return True



    def show_pst_good_format(self,final_pst):
        head = final_pst[0]
        body = final_pst[1::]
        for i in range(0,len(head)):
            print('{0:<10}'.format(head[i]), end="")
        print("")
        for i in range(0,len(body)):
            for j in range(0,len(body[i])):
                print('{0:<10}'.format(body[i][j]), end="")
            print("")





    def remove_duplicate_pst(self,pst_table):
        final_pst = []
        pst_dict = {}
        for i in range(0,len(pst_table)):
            one_pst = ''.join(pst_table[i])
            if one_pst not in pst_dict.keys():
                pst_dict[one_pst] = 1
                final_pst.append(pst_table[i])
            else:
                None
        return final_pst








if __name__ == '__main__':
    test = DEAL_XLSX('./top_pst.xlsx')
    test.read_excel_xlsx('Sheet1')
    test.show_xlxs_content()
    test.get_supply_net_info()
    # print(test.pst_table)
    # print(test.dvfs)
    # print(test.same_net)
    # print(test.all_net)
    #print(test.ban_pst_table)
    # print(test.same_net_dict)
    # print(test.all_net_dict)
    # print(test.dvfs_dict)
    print("============================================")
    test.gen_final_pst_table()
    #print(test.final_pst_table)
    test.show_pst_good_format(test.final_pst_table)
